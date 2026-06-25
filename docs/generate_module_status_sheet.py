"""Generate Ziplofy Shopify Module Status Tracker for Google Sheets import."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPORT_DATE = date(2026, 6, 22)
OUTPUT = Path(__file__).parent / "Ziplofy_Shopify_Module_Status_Tracker.xlsx"

STATUS_COLORS = {
    "Complete": "C6EFCE",
    "Partial": "FFEB9C",
    "Stub": "FCE4D6",
    "Missing": "FFC7CE",
}

MODULES = [
    # Tier, Tier Name, Category, Module, Status, Completion %, Current State / Notes, Suggested Priority
    (1, "Commerce Core", "Catalog & Inventory", "Products & variants", "Complete", 95, "Full CRUD, search, soft delete", "Maintain"),
    (1, "Commerce Core", "Catalog & Inventory", "Collections", "Complete", 90, "Theme templates recently added", "Maintain"),
    (1, "Commerce Core", "Catalog & Inventory", "Inventory levels", "Complete", 90, "Per-location management", "Maintain"),
    (1, "Commerce Core", "Catalog & Inventory", "Gift cards", "Complete", 85, "Admin + timeline", "Maintain"),
    (1, "Commerce Core", "Supply Chain", "Purchase orders", "Complete", 90, "Create, receive, tags", "Maintain"),
    (1, "Commerce Core", "Supply Chain", "Transfers & shipments", "Complete", 90, "Warehouse-style ops", "Maintain"),
    (1, "Commerce Core", "Supply Chain", "Vendors", "Complete", 85, "Vendor management operational", "Maintain"),
    (1, "Commerce Core", "Pricing & Promotions", "Discounts (4 types)", "Complete", 95, "Best-in-class; storefront validation", "Maintain"),
    (1, "Commerce Core", "Customers", "Customers", "Complete", 85, "List, detail, addresses, timeline", "Maintain"),
    (1, "Commerce Core", "Customers", "Customer segments", "Complete", 85, "Segment CRUD and entries", "Maintain"),
    (1, "Commerce Core", "Customers", "Tag management", "Complete", 90, "Product/customer/PO/transfer tags", "Maintain"),
    (1, "Commerce Core", "Recovery", "Abandoned carts (view)", "Complete", 75, "List + recovery UI; automation weak", "Medium"),
    (2, "Orders & Checkout", "Orders", "Orders list & detail", "Complete", 80, "Read-only admin", "High"),
    (2, "Orders & Checkout", "Orders", "Create order / drafts", "Stub", 25, "UI exists; no create-order API", "Critical"),
    (2, "Orders & Checkout", "Orders", "Order fulfillment", "Missing", 10, "No fulfill/cancel/refund/edit", "Critical"),
    (2, "Orders & Checkout", "Orders", "Returns & refunds", "Stub", 15, "Policy text only; templates exist, no workflow", "Critical"),
    (2, "Orders & Checkout", "Checkout", "Storefront checkout", "Partial", 50, "Creates orders; client-trusted totals", "Critical"),
    (2, "Orders & Checkout", "Checkout", "Checkout settings", "Complete", 90, "Large settings surface", "Maintain"),
    (2, "Orders & Checkout", "Payments", "Manual payments (UPI/bank/COD)", "Complete", 85, "Recently built", "Maintain"),
    (2, "Orders & Checkout", "Payments", "Payment gateways (Stripe/Razorpay)", "Stub", 10, "Provider catalog; no OAuth/capture", "Critical"),
    (2, "Orders & Checkout", "Payments", "Transactions", "Complete", 75, "Manual confirmations only", "High"),
    (3, "Shipping, Tax & Markets", "Shipping", "Shipping profiles/zones/rates", "Complete", 90, "Deep configuration", "Maintain"),
    (3, "Shipping, Tax & Markets", "Shipping", "Local delivery", "Partial", 60, "Some API wiring gaps", "High"),
    (3, "Shipping, Tax & Markets", "Shipping", "Order-linked shipping labels", "Missing", 5, "Transfers only, not orders", "High"),
    (3, "Shipping, Tax & Markets", "Tax", "Taxes & duties (config)", "Complete", 85, "India tax detail page ~1600 lines", "Maintain"),
    (3, "Shipping, Tax & Markets", "Tax", "Tax at checkout", "Partial", 40, "Not recalculated server-side", "Critical"),
    (3, "Shipping, Tax & Markets", "Markets", "Markets", "Complete", 75, "Admin CRUD", "Medium"),
    (3, "Shipping, Tax & Markets", "Markets", "Market catalogs", "Complete", 80, "Catalog market assignments", "Medium"),
    (3, "Shipping, Tax & Markets", "Markets", "Multi-currency checkout", "Missing", 15, "Config exists; checkout doesn't use it", "High"),
    (3, "Shipping, Tax & Markets", "Shipping", "Free shipping at checkout", "Partial", 50, "Static rate TODO in backend", "High"),
    (4, "Online Store & Themes", "Themes", "Theme library & install", "Complete", 90, "Install and manage themes", "Maintain"),
    (4, "Online Store & Themes", "Themes", "Theme editor / code editor", "Complete", 85, "In-browser theme editing", "Maintain"),
    (4, "Online Store & Themes", "Themes", "Custom theme builder", "Complete", 80, "444-file subsystem", "Maintain"),
    (4, "Online Store & Themes", "Themes", "Storefront SSR render", "Complete", 75, "Liquid + React packs", "Medium"),
    (4, "Online Store & Themes", "Content", "Store menus", "Complete", 85, "Navigation menu management", "Maintain"),
    (4, "Online Store & Themes", "Content", "Files / media", "Complete", 80, "S3 uploads", "Maintain"),
    (4, "Online Store & Themes", "Content", "Online store pages", "Stub", 20, "Hardcoded placeholders", "Medium"),
    (4, "Online Store & Themes", "Content", "Blog / articles", "Stub", 15, "Placeholder HTML in render controller", "Medium"),
    (4, "Online Store & Themes", "Content", "URL redirects", "Stub", 10, "Empty UI", "Low"),
    (4, "Online Store & Themes", "Content", "Metaobjects / metafields", "Stub", 15, "Coming soon in settings", "Medium"),
    (4, "Online Store & Themes", "Policies", "Store policies", "Complete", 85, "Terms, privacy, shipping, returns", "Maintain"),
    (4, "Online Store & Themes", "Accounts", "Customer accounts (storefront)", "Partial", 55, "Auth works; social login partial", "High"),
    (5, "Marketing & Growth", "Marketing", "Marketing overview", "Partial", 40, "Presentational stats", "Medium"),
    (5, "Marketing & Growth", "Marketing", "Campaigns", "Stub", 15, "Empty handlers", "Medium"),
    (5, "Marketing & Growth", "Marketing", "Attribution", "Stub", 10, "Always No data", "Low"),
    (5, "Marketing & Growth", "Automation", "Automations (UI)", "Partial", 55, "Large builder; no execution engine", "High"),
    (5, "Marketing & Growth", "Tracking", "Pixels / customer events", "Complete", 75, "Pixel management operational", "Maintain"),
    (5, "Marketing & Growth", "Merchandising", "Product offers (PDP badges)", "Complete", 80, "PDP badge configuration", "Maintain"),
    (6, "Analytics & Reporting", "Dashboard", "Dashboard / home", "Partial", 35, "Mock chart data", "High"),
    (6, "Analytics & Reporting", "Dashboard", "Analytics page", "Partial", 35, "Reuses mock dashboard", "High"),
    (6, "Analytics & Reporting", "Reports", "Reports", "Missing", 0, "Sidebar link; no route", "High"),
    (6, "Analytics & Reporting", "Reports", "Live view", "Missing", 0, "Sidebar link; no route", "Medium"),
    (6, "Analytics & Reporting", "Exports", "Export logs", "Missing", 5, "Backend route not mounted", "Medium"),
    (7, "Settings & Platform Admin", "Store", "General settings", "Complete", 85, "Core store configuration", "Maintain"),
    (7, "Settings & Platform Admin", "Store", "Branding", "Complete", 85, "Logo, colors, identity", "Maintain"),
    (7, "Settings & Platform Admin", "Access", "Roles & permissions", "Complete", 80, "Permission tree + API", "Maintain"),
    (7, "Settings & Platform Admin", "Access", "Users / staff", "Partial", 65, "Staff management gaps", "Medium"),
    (7, "Settings & Platform Admin", "Store", "Locations", "Complete", 85, "Location CRUD", "Maintain"),
    (7, "Settings & Platform Admin", "Notifications", "Notifications", "Partial", 60, "Test email/toggle TODOs", "Medium"),
    (7, "Settings & Platform Admin", "Integrations", "Webhooks", "Partial", 30, "Static signing key; create flow TODO", "High"),
    (7, "Settings & Platform Admin", "Domains", "Domains", "Partial", 25, "Mock data", "High"),
    (7, "Settings & Platform Admin", "Localization", "Languages", "Stub", 5, "Coming soon", "Low"),
    (7, "Settings & Platform Admin", "Integrations", "Apps & sales channels", "Stub", 5, "Placeholder", "Medium"),
    (7, "Settings & Platform Admin", "Billing", "Plan / billing (merchant SaaS)", "Partial", 45, "Membership plans hook; billing UI scaffold", "High"),
    (7, "Settings & Platform Admin", "Privacy", "Customer privacy", "Partial", 50, "Privacy settings partial", "Medium"),
    (7, "Settings & Platform Admin", "Audit", "Activity log", "Partial", 30, "Mock entries", "Medium"),
    (7, "Settings & Platform Admin", "Security", "API security (auth on routes)", "Partial", 50, "Many admin routes lack protect", "Critical"),
    (8, "Infrastructure & Platform", "Auth", "Admin auth (OTP/JWT)", "Complete", 85, "OTP + JWT authentication", "Maintain"),
    (8, "Infrastructure & Platform", "Auth", "Storefront customer auth", "Complete", 80, "Customer login/register", "Maintain"),
    (8, "Infrastructure & Platform", "Email", "Email templates", "Complete", 70, "Order confirm works", "Medium"),
    (8, "Infrastructure & Platform", "Email", "Email queue (BullMQ)", "Missing", 10, "Disabled in index.ts", "High"),
    (8, "Infrastructure & Platform", "Integrations", "Webhooks for apps", "Missing", 0, "Not implemented", "High"),
    (8, "Infrastructure & Platform", "Integrations", "App marketplace / API keys", "Missing", 0, "Not implemented", "Medium"),
    (8, "Infrastructure & Platform", "Automation", "Background automation runner", "Missing", 5, "Flows stored, not executed", "High"),
]

TIER_SUMMARY = [
    (1, "Commerce Core", 12, 12, 88, "Strongest area — catalog, discounts, inventory ops"),
    (2, "Orders & Checkout", 9, 4, 49, "Biggest gap — fulfillment, gateways, order lifecycle"),
    (3, "Shipping, Tax & Markets", 9, 6, 61, "Config strong; checkout tax/currency gaps"),
    (4, "Online Store & Themes", 12, 8, 58, "Differentiator — theme builder strong; CMS weak"),
    (5, "Marketing & Growth", 6, 2, 46, "Deferred to Phase 2 — UI present; execution engine missing"),
    (6, "Analytics & Reporting", 5, 0, 15, "Deferred to Phase 2 — mostly mock data or missing routes"),
    (7, "Settings & Platform Admin", 14, 5, 52, "Core settings done; activity log deferred to Phase 2"),
    (8, "Infrastructure & Platform", 7, 3, 41, "Auth solid; queues, webhooks, automation missing"),
]

PLATFORM_SCALE = [
    ("Admin routes", "~120 routes in App.tsx"),
    ("Admin pages", "154+ page components"),
    ("React contexts", "~99 (heavy context-driven architecture)"),
    ("Backend API prefixes", "~115 under /api/*"),
    ("Controllers", "123"),
    ("Mongoose models", "143+"),
    ("Theme builder", "444+ files in create-theme/"),
]

LEGEND = [
    ("Complete", "End-to-end (UI + API + real data)"),
    ("Partial", "Works but gaps, mocks, or missing backend"),
    ("Stub", "UI shell / Coming soon"),
    ("Missing", "Not built or not routed"),
]

# Modules deferred to Phase 2 — marketing, analytics, and activity log
PHASE_2_MODULE_NAMES = {
    "Marketing overview",
    "Campaigns",
    "Attribution",
    "Automations (UI)",
    "Pixels / customer events",
    "Product offers (PDP badges)",
    "Dashboard / home",
    "Analytics page",
    "Reports",
    "Live view",
    "Export logs",
    "Activity log",
}

PHASE_FILL = {
    "Phase 1": PatternFill("solid", fgColor="E8F4FD"),
    "Phase 2": PatternFill("solid", fgColor="F3E5F5"),
}

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E79")
wrap = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def module_phase(mod: str) -> str:
    return "Phase 2" if mod in PHASE_2_MODULE_NAMES else "Phase 1"


def phase2_modules():
    return [m for m in MODULES if module_phase(m[3]) == "Phase 2"]


def write_module_rows(ws, modules, start_row, headers):
    for r, row in enumerate(modules, start_row):
        tier, tier_name, cat, mod, status, pct, notes, priority = row
        phase = module_phase(mod)
        values = [
            phase,
            tier,
            tier_name,
            cat,
            mod,
            status,
            pct / 100,
            notes,
            priority,
            "",
            "",
            "",
            "",
            "",
            REPORT_DATE,
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = wrap
            if c == 1:
                cell.fill = PHASE_FILL[phase]
                cell.font = Font(bold=True)
            if c == 6 and status in STATUS_COLORS:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
            if c == 7:
                cell.number_format = "0%"


def auto_width(ws, min_width=10, max_width=45):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def build_tracker(wb):
    ws = wb.active
    ws.title = "Module Tracker"
    ws["A1"] = "Ziplofy — Shopify Module Status Tracker"
    ws["A1"].font = title_font
    ws.merge_cells("A1:O1")
    ws["A2"] = (
        f"Report snapshot date: {REPORT_DATE.strftime('%d %b %Y')}  |  "
        "Phase 1 = launch scope  |  Phase 2 = marketing, analytics & activity log (deferred)"
    )
    ws.merge_cells("A2:O2")
    ws["A2"].font = Font(italic=True, color="666666")

    headers = [
        "Phase",
        "Tier",
        "Tier Name",
        "Category",
        "Module",
        "Status",
        "Completion %",
        "Current State / Notes",
        "Suggested Priority",
        "Date Started",
        "Target End Date",
        "Date Completed",
        "Owner",
        "Comments",
        "Report Date",
    ]
    start = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, len(headers))

    phase1 = [m for m in MODULES if module_phase(m[3]) == "Phase 1"]
    phase2 = phase2_modules()
    row = start + 1

    ws.cell(row=row, column=1, value="PHASE 1 — LAUNCH SCOPE").font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    row += 1
    write_module_rows(ws, phase1, row, headers)
    row += len(phase1) + 1

    ws.cell(row=row, column=1, value="PHASE 2 — MARKETING, ANALYTICS & AUDIT").font = Font(
        bold=True, size=12, color="6A1B9A"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    row += 1
    write_module_rows(ws, phase2, row, headers)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{start}:O{row + len(phase2) - 1}"
    auto_width(ws)


def build_phase2_roadmap(wb):
    ws = wb.create_sheet("Phase 2 Roadmap")
    ws["A1"] = "Phase 2 Roadmap — Marketing, Analytics & Activity Log"
    ws["A1"].font = title_font
    ws.merge_cells("A1:O1")
    ws["A2"] = (
        "Deferred post-launch scope. Includes all Tier 5 (Marketing & Growth), "
        "Tier 6 (Analytics & Reporting), and Activity log from Settings."
    )
    ws.merge_cells("A2:O2")
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = f"Modules in Phase 2: {len(phase2_modules())}  |  Fill Date Started / Target End Date for Phase 2 sprint planning"
    ws.merge_cells("A3:O3")
    ws["A3"].font = Font(italic=True, color="666666")

    headers = [
        "Phase",
        "Tier",
        "Tier Name",
        "Category",
        "Module",
        "Status",
        "Completion %",
        "Current State / Notes",
        "Suggested Priority",
        "Date Started",
        "Target End Date",
        "Date Completed",
        "Owner",
        "Comments",
        "Report Date",
    ]
    start = 5
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, len(headers))

    sections = [
        ("Marketing & Growth (Tier 5)", [m for m in phase2_modules() if m[1] == "Marketing & Growth"]),
        ("Analytics & Reporting (Tier 6)", [m for m in phase2_modules() if m[1] == "Analytics & Reporting"]),
        ("Audit & Compliance", [m for m in phase2_modules() if m[3] == "Activity log"]),
    ]

    row = start + 1
    for section_title, items in sections:
        ws.cell(row=row, column=1, value=section_title).font = Font(bold=True, color="6A1B9A")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1
        write_module_rows(ws, items, row, headers)
        row += len(items) + 1

    p2 = phase2_modules()
    # summary row
    summary_row = row
    ws.cell(row=summary_row, column=1, value="Phase 2 summary").font = Font(bold=True)
    avg_pct = sum(m[5] for m in p2) / len(p2) / 100
    ws.cell(row=summary_row, column=7, value=avg_pct).number_format = "0%"
    ws.cell(row=summary_row, column=8, value=f"{len(p2)} modules — avg completion at handoff")

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{start}:O{row - 2}"
    auto_width(ws)


def build_tier_summary(wb):
    ws = wb.create_sheet("Tier Summary")
    ws["A1"] = "Tier Scorecard"
    ws["A1"].font = title_font
    headers = [
        "Tier",
        "Tier Name",
        "Total Modules",
        "Complete Modules",
        "Avg Completion %",
        "Assessment",
        "Date Started",
        "Target End Date",
        "Date Completed",
        "Owner",
        "Comments",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))
    for r, row in enumerate(TIER_SUMMARY, 4):
        for c, val in enumerate(row, 1):
            val_out = val / 100 if c == 5 else val
            cell = ws.cell(row=r, column=c, value=val_out)
            cell.border = border
            cell.alignment = wrap
            if c == 5:
                cell.number_format = "0%"
        for c in range(7, 12):
            ws.cell(row=r, column=c, value="").border = border
    ws["A13"] = "Overall platform assessment"
    ws["A13"].font = Font(bold=True)
    ws.merge_cells("A13:K13")
    ws["A14"] = (
        "The product already feels like Shopify in catalog, discounts, shipping/tax config, "
        "themes, and inventory ops. It is not yet at parity for payments, order lifecycle, "
        "fulfillment, returns, analytics, CMS, and app ecosystem."
    )
    ws.merge_cells("A14:K14")
    ws["A14"].alignment = wrap
    auto_width(ws)


def build_platform_scale(wb):
    ws = wb.create_sheet("Platform Scale")
    ws["A1"] = "Platform Scale Metrics"
    ws["A1"].font = title_font
    headers = ["Metric", "Value", "Comments"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 3)
    for r, (metric, value) in enumerate(PLATFORM_SCALE, 4):
        ws.cell(row=r, column=1, value=metric).border = border
        ws.cell(row=r, column=2, value=value).border = border
        ws.cell(row=r, column=3, value="").border = border
    auto_width(ws)


def build_legend(wb):
    ws = wb.create_sheet("Legend")
    ws["A1"] = "Status Legend"
    ws["A1"].font = title_font
    headers = ["Status", "Definition"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 2)
    for r, (status, definition) in enumerate(LEGEND, 4):
        ws.cell(row=r, column=1, value=status).border = border
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
        ws.cell(row=r, column=2, value=definition).border = border
    auto_width(ws)


def main():
    wb = Workbook()
    build_tracker(wb)
    build_phase2_roadmap(wb)
    build_tier_summary(wb)
    build_platform_scale(wb)
    build_legend(wb)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
